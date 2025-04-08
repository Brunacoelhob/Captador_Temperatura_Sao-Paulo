import tkinter as tk
from tkinter import messagebox
import requests
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import matplotlib.pyplot as plt

usuario_nome = ""
usuario_email = ""

def pegar_dados_clima():
    try:
        chave_api = "44cef0e07f922c7b235c291d9c1f4308"
        cidade = "Sao Paulo,BR"
        url = f"https://api.openweathermap.org/data/2.5/weather?q={cidade}&appid={chave_api}&units=metric&lang=pt_br"
        resposta = requests.get(url)
        dados = resposta.json()
        temperatura = dados["main"]["temp"]
        umidade = dados["main"]["humidity"]
        return temperatura, umidade
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao obter dados da API: {e}")
        return None, None

def salvar_em_planilha(temperatura, umidade):
    arquivo = "dados_clima.xlsx"
    data_hora = datetime.now().strftime("%d/%m/%Y %H:%M")
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Data/Hora", "Temperatura (°C)", "Umidade (%)", "Nome", "E-mail"])

    ws.append([data_hora, temperatura, umidade, usuario_nome, usuario_email])
    wb.save(arquivo)

def exibir_previsao_personalizada(temperaturas, umidades):
    dias_semana = ["Seg", "Ter", "Qua", "Qui", "Sex"]
    janela_prev = tk.Toplevel(janela)
    janela_prev.title("Previsão do Tempo")
    janela_prev.geometry("400x350")
    janela_prev.configure(bg="white")

    titulo = tk.Label(janela_prev, text="Previsão do Tempo", font=("Arial", 16, "bold"), bg="#2c5282", fg="white", pady=10)
    titulo.pack(fill="x")

    subtitulo = tk.Label(janela_prev, text="São Paulo - SP", font=("Arial", 12), bg="#2c5282", fg="white")
    subtitulo.pack(fill="x")

    tabela = tk.Frame(janela_prev, bg="white")
    tabela.pack(pady=20)

    headers = ["Dia", "Temperatura", "Umidade"]
    for col, texto in enumerate(headers):
        tk.Label(tabela, text=texto, font=("Arial", 11, "bold"), bg="#03A9F4", fg="white", width=12).grid(row=0, column=col, padx=5, pady=5)

    for i in range(5):
        tk.Label(tabela, text=dias_semana[i], font=("Arial", 10), bg="white", width=12).grid(row=i+1, column=0, pady=2)
        tk.Label(tabela, text=temperaturas[i], font=("Arial", 10), bg="white", width=12).grid(row=i+1, column=1, pady=2)
        tk.Label(tabela, text=umidades[i], font=("Arial", 10), bg="white", width=12).grid(row=i+1, column=2, pady=2)

    tk.Button(janela_prev, text="OK", command=janela_prev.destroy, bg="#4caf50", fg="white",
              font=("Arial", 10, "bold"), width=10, relief="flat").pack(pady=10)

def buscar_previsao():
    temperatura, umidade = pegar_dados_clima()
    if temperatura is not None and umidade is not None:
        salvar_em_planilha(temperatura, umidade)
        temperaturas = ["28°C", "23°C", "18°C", "16°C", f"{temperatura:.1f}°C"]
        umidades = ["32%", "46%", "58%", "70%", f"{umidade}%"]
        exibir_previsao_personalizada(temperaturas, umidades)

def visualizar_dados():
    if not os.path.exists("dados_clima.xlsx"):
        messagebox.showwarning("Aviso", "Nenhum dado encontrado.")
        return

    wb = load_workbook("dados_clima.xlsx")
    ws = wb.active
    todas_linhas = list(ws.iter_rows(min_row=2, values_only=True))
    ultimos = todas_linhas[-10:]

    nova_janela = tk.Toplevel(janela)
    nova_janela.title("Dados Salvos")
    nova_janela.geometry("900x500")
    nova_janela.configure(bg="white")

    header = tk.Frame(nova_janela, bg="#2c5282", height=70)
    header.pack(fill="x")

    tk.Label(header, text="Dados Salvos", font=("Arial", 16, "bold"), bg="#2c5282", fg="white").pack(pady=(10, 0))
    tk.Label(header, text="Histórico da Previsão", font=("Arial", 12), bg="#2c5282", fg="white").pack()

    tabela = tk.Frame(nova_janela, bg="white")
    tabela.pack(pady=20)

    headers = ["Data/Hora", "Temperatura (°C)", "Umidade (%)", "Nome", "E-mail"]
    for col, texto in enumerate(headers):
        tk.Label(tabela, text=texto, font=("Arial", 10, "bold"), bg="#03A9F4", fg="white", width=20).grid(row=0, column=col)

    for i, row in enumerate(ultimos):
        for j, val in enumerate(row):
            txt = f"{int(val)}%" if j == 2 else str(val)
            if j == 1:
                txt = f"{int(float(val))}°C"
            tk.Label(tabela, text=txt, font=("Arial", 10), bg="white", width=20).grid(row=i+1, column=j)

    tk.Button(nova_janela, text="OK", command=nova_janela.destroy, bg="#4caf50", fg="white",
              font=("Arial", 10, "bold"), width=10, relief="flat").pack(pady=10)

def gerar_grafico():
    if not os.path.exists("dados_clima.xlsx"):
        messagebox.showerror("Erro", "Nenhuma planilha encontrada.")
        return

    wb = load_workbook("dados_clima.xlsx")
    ws = wb.active
    linhas = list(ws.iter_rows(min_row=2, values_only=True))[-10:]

    datas = [linha[0].replace(" ", "\n") for linha in linhas]
    temperaturas = [round(float(linha[1])) for linha in linhas]
    umidades = [linha[2] for linha in linhas]

    plt.figure(figsize=(12, 5))
    plt.plot(datas, temperaturas, label="Temperatura (°C)", marker='o', linewidth=2, color='royalblue')
    plt.plot(datas, umidades, label="Umidade (%)", marker='o', linewidth=2, color='gray')

    plt.suptitle("Gráfico de Variação - Temperatura e Umidade\nSão Paulo - SP", fontsize=14, fontweight='bold', color='white',
                 bbox=dict(facecolor='#2c5282', edgecolor='none', boxstyle='round,pad=0.5'))

    plt.grid(True, linestyle='--', alpha=0.3)
    plt.xlabel("Data e Hora")
    plt.ylabel("Valores")
    plt.legend()
    plt.tight_layout(rect=[0, 0, 1, 0.93])
    plt.show()

def abrir_planilha():
    if os.path.exists("dados_clima.xlsx"):
        os.startfile("dados_clima.xlsx")
    else:
        messagebox.showwarning("Aviso", "Planilha ainda não foi criada.")

def tela_principal():
    for widget in janela.winfo_children():
        widget.destroy()

    janela.geometry("600x500")
    janela.configure(bg="white")

    header = tk.Frame(janela, bg="#2c5282", height=70)
    header.pack(fill="x")

    tk.Label(header, text="Previsão do Tempo", font=("Arial", 18, "bold"), bg="#2c5282", fg="white").pack(pady=(10, 0))
    tk.Label(header, text="São Paulo - SP", font=("Arial", 12), bg="#2c5282", fg="white").pack()

    botoes = tk.Frame(janela, bg="white")
    botoes.pack(pady=30)

    estilo = {"font": ("Arial", 14, "bold"), "width": 20, "height": 2, "bd": 0}

    tk.Button(botoes, text="Buscar Previsão", command=buscar_previsao, bg="#FFD600", fg="white", **estilo).pack(pady=5)
    tk.Button(botoes, text="Visualizar Dados", command=visualizar_dados, bg="#388E3C", fg="white", **estilo).pack(pady=5)
    tk.Button(botoes, text="Gerar Gráfico", command=gerar_grafico, bg="#0288D1", fg="white", **estilo).pack(pady=5)
    tk.Button(botoes, text="Visualizar Planilha", command=abrir_planilha, bg="#9C27B0", fg="white", **estilo).pack(pady=5)
    tk.Button(janela, text="Voltar", command=tela_inicial, bg="gray", fg="white",
              font=("Arial", 10, "bold"), width=10, relief="flat").pack(pady=10)

def iniciar():
    global usuario_nome, usuario_email
    usuario_nome = entry_nome.get()
    usuario_email = entry_email.get()
    if not usuario_nome or not usuario_email:
        messagebox.showwarning("Atenção", "Preencha todos os campos.")
        return
    tela_principal()

def tela_inicial():
    for widget in janela.winfo_children():
        widget.destroy()

    janela.geometry("500x300")
    janela.configure(bg="white")

    header = tk.Frame(janela, bg="#2c5282", height=60)
    header.pack(fill="x")

    tk.Label(header, text="Login", font=("Arial", 18, "bold"), bg="#2c5282", fg="white").pack(pady=10)

    conteudo = tk.Frame(janela, bg="white")
    conteudo.pack(pady=30)

    global entry_nome, entry_email

    tk.Label(conteudo, text="Nome", anchor="w", font=("Arial", 12), bg="white").pack(fill="x")
    entry_nome = tk.Entry(conteudo, font=("Arial", 12), width=40, relief="solid")
    entry_nome.pack(pady=5)

    tk.Label(conteudo, text="E-mail", anchor="w", font=("Arial", 12), bg="white").pack(fill="x")
    entry_email = tk.Entry(conteudo, font=("Arial", 12), width=40, relief="solid")
    entry_email.pack(pady=5)

    tk.Button(conteudo, text="Iniciar", command=iniciar, bg="#2c5282", fg="white",
              font=("Arial", 12, "bold"), width=15, height=1, relief="flat").pack(pady=15)

# ====== INICIAR APLICAÇÃO ======
janela = tk.Tk()
janela.title("Início")
tela_inicial()
janela.mainloop()